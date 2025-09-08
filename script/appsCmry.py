import re
import os
import pickle
import pandas as pd
import numpy as np
import streamlit as st
import math
from PIL import Image
from tensorflow.keras.models import load_model
from tensorflow.keras.preprocessing.sequence import pad_sequences
from openpyxl.styles import PatternFill

script_dir = os.path.dirname(os.path.abspath(__file__))

# ============================================================
# --- CLEANING ---
# ============================================================
def clean_text(text: str) -> str:
    text = str(text)
    text = re.sub(r"^(CMI|BMI|PMI|IMI|CML|CM|PM|BM|IM|APD|CGI|FUEL|OPR|RGI)[-\s]*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"[^a-zA-Z0-9]", " ", text)   # hapus simbol selain huruf/angka
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text

# ============================================================
# --- KATEGORI ---
# ============================================================
CATEGORIES = ["CM", "PM", "BM", "IM", "UNK", "APD", "CGI", "FUEL", "OPR", "Other", "CML", "RGI"]

def extract_category(text: str) -> str:
    if not text or not isinstance(text, str):
        return "UNK"
    text = text.upper().strip()

    if re.match(r"^CML[-\s0-9A-Z]*", text):
        return "CML"
    elif re.match(r"^CM(?!L)[-\s0-9A-Z]*", text):
        return "CM"

    for cat in CATEGORIES:
        if cat in ["CM", "CML"]:
            continue
        if re.match(rf"^{cat}[-\s0-9A-Z]*", text):
            return cat

    return "UNK"

# ============================================================
# --- LOAD MODEL & ENCODERS ---
# ============================================================
@st.cache_resource
def load_assets():
    model = load_model(os.path.join(script_dir, "best_text_classifier.h5"))
    with open(os.path.join(script_dir, "tokenizer.pkl"), "rb") as f: tokenizer = pickle.load(f)
    with open(os.path.join(script_dir, "le_machine.pkl"), "rb") as f: le_machine = pickle.load(f)
    with open(os.path.join(script_dir, "le_category.pkl"), "rb") as f: le_category = pickle.load(f)
    return model, tokenizer, le_machine, le_category

model, tokenizer, le_machine, le_category = load_assets()
max_len = 50

# ============================================================
# --- HAPUS PASANGAN 201 & 202 ---
# ============================================================
def remove_201_202_pairs(df: pd.DataFrame):
    df = df.copy()
    df["Movement type"] = df["Movement type"].astype(str).str.extract(r"(\d+)")[0]
    df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0)
    df["Material"] = df["Material"].astype(str).str.strip()
    df["Material Description"] = df["Material Description"].astype(str).str.strip().str.upper()

    mask_201_202 = df["Movement type"].isin(["201", "202"])
    subset = df.loc[mask_201_202].copy()

    subset["pair_key"] = (
        subset["Material"] + "|" +
        subset["Material Description"] + "|" +
        subset["Quantity"].abs().astype(str)
    )

    keys_to_remove = subset.groupby("pair_key")["Movement type"].nunique()
    keys_to_remove = keys_to_remove[keys_to_remove == 2].index

    mask_drop = mask_201_202 & subset["pair_key"].isin(keys_to_remove)
    n_removed = mask_drop.sum()

    df = df[~mask_drop]
    return df, n_removed

# ============================================================
# --- SAFETY STOCK & ROP ---
# ============================================================
def compute_ss_rop(df,
                   date_col="Posting Date",
                   material_col="Material",
                   qty_col="Quantity",
                   leadtime_col="Lead Time",
                   leadtime_std_col="LeadTime_Std",
                   service_level=0.95,
                   min_history_days=90,
                   resample_rule="D",
                   moq=None):
    df2 = df.copy()
    df2[date_col] = pd.to_datetime(df2[date_col], errors="coerce")
    df2["__cons"] = df2[qty_col].astype(float).abs()

    mats = df2[material_col].unique()
    rows = []

    z_table = {0.90: 1.2816, 0.95: 1.645, 0.99: 2.3263}
    z = z_table.get(round(service_level, 2), 1.645)

    for m in mats:
        sub = df2[df2[material_col] == m].copy()
        if sub.empty:
            continue

        # --- Cek kalau semua tanggal null ---
        if sub[date_col].isna().all():
            rows.append({
                material_col: m,
                "history_days": 0,
                "mean_daily": np.nan,
                "std_daily": np.nan,
                "leadtime_days": np.nan,
                "leadtime_std": np.nan,
                "Z": z,
                "safety_stock": np.nan,
                "rop": np.nan,
                "note": "no_valid_date"
            })
            continue

        start = sub[date_col].min()
        end = sub[date_col].max()
        if pd.isna(start) or pd.isna(end):
            rows.append({
                material_col: m,
                "history_days": 0,
                "mean_daily": np.nan,
                "std_daily": np.nan,
                "leadtime_days": np.nan,
                "leadtime_std": np.nan,
                "Z": z,
                "safety_stock": np.nan,
                "rop": np.nan,
                "note": "invalid_date_range"
            })
            continue

        # --- Daily consumption ---
        idx = pd.date_range(start=start, end=end, freq="D")
        daily = (sub.set_index(date_col)
                   .resample("D")["__cons"]
                   .sum()
                   .reindex(idx, fill_value=0))

        history_days = (end - start).days + 1
        mean_d = daily.mean()
        std_d = daily.std(ddof=0)


        lt_val = sub[leadtime_col].dropna().unique()
        lt_val = float(lt_val[0]) if len(lt_val) > 0 else np.nan

        lt_std_val = None
        if leadtime_std_col in sub.columns:
            lt_std = sub[leadtime_std_col].dropna().unique()
            if len(lt_std) > 0:
                lt_std_val = float(lt_std[0])

        note = []
        if history_days < min_history_days:
            note.append(f"history_short ({history_days} days)")
        if np.isnan(lt_val):
            note.append("no_leadtime")

        if not math.isnan(lt_val):
            if lt_std_val:
                sigma_lt_demand = math.sqrt((mean_d**2 * lt_std_val**2) + (lt_val * std_d**2))
            else:
                sigma_lt_demand = std_d * math.sqrt(lt_val)

            ss = z * sigma_lt_demand
            rop = mean_d * lt_val + ss

            if moq and moq > 1:
                ss = math.ceil(ss / moq) * moq
                rop = math.ceil(rop / moq) * moq
        else:
            ss, rop = np.nan, np.nan

        rows.append({
            material_col: m,
            "history_days": history_days,
            "mean_daily": mean_d,
            "std_daily": std_d,
            "leadtime_days": lt_val,
            "leadtime_std": lt_std_val,
            "Z": z,
            "safety_stock": ss,
            "rop": rop,
            "note": ";".join(note) if note else ""
        })

    return pd.DataFrame(rows)

# ============================================================
# --- REORDER FORECAST ---
# ============================================================
def compute_reorder_forecast(usage_df, ss_df, stock_df,
                             material_col="Material",
                             stock_col="Unrestricted",
                             date_col="Posting Date"):
    usage_df = usage_df.copy()
    usage_df[date_col] = pd.to_datetime(usage_df[date_col], errors="coerce")
    usage_df["__cons"] = usage_df["Quantity"].astype(float).abs()

    avg_usage = usage_df.groupby(material_col)["__cons"].mean().reset_index(name="avg_daily_usage")

    merged = ss_df.merge(avg_usage, on=material_col, how="left")
    merged = merged.merge(stock_df[[material_col, stock_col]], on=material_col, how="left")

    merged["days_to_rop"] = (merged[stock_col] - merged["rop"]) / merged["avg_daily_usage"]
    merged["days_to_rop"] = merged["days_to_rop"].replace([np.inf, -np.inf], np.nan)

    merged["reorder_date"] = pd.to_datetime("today") + pd.to_timedelta(merged["days_to_rop"], unit="D")
    
    def check_alert(row):
        stock = row[stock_col]
        ss = row["safety_stock"]
        rop = row["rop"]
        try:
            stock_val = pd.to_numeric(stock, errors="coerce")
        except:
            stock_val = None

        if stock is None or pd.isna(stock_val) or str(stock).strip() == "" or stock_val <= 0:
            return "Out of Stock"
        if pd.isna(rop) or pd.isna(ss):
            return "No ROP Data"
        if stock_val < ss:
            return "CRITICAL – Below Safety Stock"
        if stock_val < rop:
            return "Warning – Below ROP"
        return "Safe – No Action Needed"


    merged["Alert"] = merged.apply(check_alert, axis=1)
    return merged[[material_col, stock_col, "avg_daily_usage", "safety_stock", "rop", "days_to_rop", "reorder_date", "Alert"]]

# ============================================================
# --- PREDIKSI DATAFRAME ---
# ============================================================
def predict_dataframe(df: pd.DataFrame,
                      service_level=0.95,
                      min_history_days=90,
                      moq=None):
    df = df.copy()
    df["clean_text"] = df["Document Header Text"].apply(clean_text)
    df["Category"] = df["Document Header Text"].apply(extract_category)

    X_seq = pad_sequences(tokenizer.texts_to_sequences(df["clean_text"]), maxlen=max_len)
    pred_machine = model.predict(X_seq)
    if isinstance(pred_machine, (list, tuple)):
        pred_machine = pred_machine[0]

    df["Machine"] = le_machine.inverse_transform(np.argmax(pred_machine, axis=1))
    df.loc[df["Category"] == "APD", "Machine"] = "APD"

    if {"Movement type", "Quantity", "Material", "Material Description"}.issubset(df.columns):
        df, n_removed = remove_201_202_pairs(df)
        if n_removed > 0:
            st.info(f"🔎 {n_removed} baris (pasangan 201 & 202) otomatis dihapus.")

    before = len(df)
    df = df[~df["Category"].isin(["CGI", "RGI"])]
    dropped = before - len(df)
    if dropped > 0:
        st.info(f"🗑️ {dropped} baris kategori CGI/RGI otomatis dihapus.")

    df = df.drop(columns=["clean_text"])

    summary = (
        df.groupby(["Machine", "Category"])
        .size()
        .reset_index(name="Count")
        .pivot(index="Machine", columns="Category", values="Count")
        .fillna(0)
        .astype(int)
    )

    ss_summary = compute_ss_rop(
        df,
        date_col="Posting Date",
        material_col="Material",
        qty_col="Quantity",
        leadtime_col="Lead Time",
        service_level=service_level,
        min_history_days=min_history_days,
        moq=moq
    )

    return df, summary, ss_summary

# ============================================================
# --- STREAMLIT UI ---
# ============================================================
st.set_page_config(
    page_title="Cimory Sparepart Management",
    layout="centered",
    initial_sidebar_state="expanded"
)

st.markdown(
    """
    <style>
    .block-container {
        max-width: 1100px;   /* atur sesuai kebutuhan */
        padding-left: 2rem;
        padding-right: 2rem;
    }
    </style>
    """,
    unsafe_allow_html=True
)

logo_path = os.path.join(script_dir, "Cimory.png")
logo = Image.open(logo_path)
st.image(logo, width=200)

st.title("Sparepart Management and Inventory Planner")

st.markdown(
    """
    <style>
    .cta-button {
        background-color: #433e8b;
        color: white !important;
        border: none;
        padding: 12px 24px;
        border-radius: 10px;
        font-size: 16px;
        font-weight: bold;
        cursor: pointer;
        text-decoration: none !important;
        display: inline-block;
        transition: background-color 0.3s ease;
    }
    .cta-button:hover {
        background-color: #e5232a;
        color: white !important;
        text-decoration: none !important;
    }
    .cta-container {
        margin-bottom: 20px; /* jarak ke elemen bawah */
    }
    </style>

    <div class="cta-container">
        <a href="https://bit.ly/Handbook-SSMCMRY" target="_blank" class="cta-button">
            Klik untuk mengakses Handbook Sparepart Management System
        </a>
    </div>
    """,
    unsafe_allow_html=True
)

uploaded_mb51 = st.file_uploader("📂 Upload MB51 (historical)", type=["csv", "xlsx"])
uploaded_mb52 = st.file_uploader("📂 Upload MB52 (current stock)", type=["csv", "xlsx"])

service_level = st.slider("🎯 Service Level", 0.8, 0.99, 0.95, 0.01)
min_history_days = st.number_input("📆 Minimal history days", min_value=30, value=90, step=30)
moq = st.number_input("📦 MOQ / Lot Size (isi 1 kalau bebas)", min_value=1, value=1, step=1)

st.markdown(
    """
    <style>
    .watermark {
        position: fixed;
        bottom: 10px;
        left: 50%;
        transform: translateX(-50%);
        font-size: 14px;
        color: grey;
        opacity: 0.7;
        z-index: 100;
    }
    </style>
    <div class="watermark">Powered by Digital Transformation CMD Sentul</div>
    """,
    unsafe_allow_html=True
)


if uploaded_mb51 and uploaded_mb52 and st.button("🚀 Jalankan Prediksi"):
    # Load MB51
    if uploaded_mb51.name.endswith(".csv"):
        df = pd.read_csv(uploaded_mb51, encoding="latin1", on_bad_lines="skip")
    else:
        df = pd.read_excel(uploaded_mb51)

    # --- Bersihin baris yang Posting Date dan Material kosong ---
    df = df.dropna(subset=["Posting Date", "Material"], how="all")
    df = df[~df["Material"].astype(str).str.strip().isin(["", "nan", "NaN", "None"])]

    # Merge dengan lead_time.xlsx
    try:
        leadtime_df = pd.read_excel(os.path.join(script_dir, "lead_time.xlsx"))
        df = df.merge(leadtime_df, on="Material", how="left")
    except Exception as e:
        st.warning(f"⚠️ Gagal load lead_time.xlsx: {e}")

    # Load MB52
    if uploaded_mb52.name.endswith(".csv"):
        stock_df = pd.read_csv(uploaded_mb52, encoding="latin1", on_bad_lines="skip")
    else:
        stock_df = pd.read_excel(uploaded_mb52)

    with st.spinner("🔄 Sedang memproses..."):
        result_df, summary, ss_summary = predict_dataframe(
            df,
            service_level=service_level,
            min_history_days=min_history_days,
            moq=moq if moq > 1 else None
        )
        reorder_df = compute_reorder_forecast(result_df, ss_summary, stock_df)

    st.success("✅ Prediksi selesai!")

    st.write("📊 Data hasil klasifikasi:")
    st.dataframe(result_df.head())
    st.write("📊 Ringkasan per Mesin & Kategori:")
    st.dataframe(summary)
    st.write("📦 Safety Stock & ROP per Material:")
    st.dataframe(ss_summary.head())
    st.write("📦 Rekomendasi Reorder:")
    st.dataframe(reorder_df.head())

    out_file = "EXPORT_RESULT.xlsx"
    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        result_df.to_excel(writer, sheet_name="Data", index=False)
        summary.to_excel(writer, sheet_name="Summary")
        ss_summary.to_excel(writer, sheet_name="SS_ROP", index=False)
        reorder_df.to_excel(writer, sheet_name="Reorder", index=False)

        # --- Terapkan warna untuk kolom Alert di sheet Reorder ---
        ws = writer.sheets["Reorder"]
        alert_col = reorder_df.columns.get_loc("Alert") + 1  # openpyxl 1-based

        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")       # merah untuk Out of Stock
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")   # oranye untuk Critical
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")   # kuning untuk Warning
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")    # hijau untuk Safe

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=alert_col, max_col=alert_col):
            for cell in row:
                if "Out of Stock" in str(cell.value):
                    cell.fill = red_fill
                elif "CRITICAL" in str(cell.value):
                    cell.fill = orange_fill
                elif "Warning" in str(cell.value):
                    cell.fill = yellow_fill
                elif "Safe" in str(cell.value):
                    cell.fill = green_fill

    with open(out_file, "rb") as f:
        st.download_button(
            "⬇️ Download XLSX (MB51+MB52)",
            f,
            file_name=out_file,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
